# 1. 각각의 엑셀데이터를 하나의 엑셀을 만들어 시트로 이동

import openpyxl # Excel 라이브러리
import os
import pandas as pd

# 1-1. 비어있는 엑셀 만들기
wb = openpyxl.Workbook()
wb.create_sheet('car', 0)
wb.create_sheet('semiconductor', 0)
wb.create_sheet('metaverse', 0)
wb.create_sheet('battery', 0)
wb.save('product_info.xlsx')

# 1-2. 'practice/참고자료' 폴더의 Excel 파일 목록 담기
path = "practice/참고자료/"

file_list = os.listdir(path)
file_list_xls = []
for file in file_list:
    if ".xlsx" in file:
        file_list_xls.append(file)

# 1-3. Excel 파일 목록의 데이터프레임 출력
for file in file_list_xls:
    df = pd.read_excel(f"practice/참고자료/{file}")
    print(df)

# 1-4. 불러온 엑셀 파일의 내용을 product_info.xls 시트에 한 개씩 넣고 저장하기
path_dir = "practice/참고자료/"
files = os.listdir(path_dir)
file_list_xls = []
for file in files:
    if '.xls' in file:
        file_list_xls.append(file)

file_nm = "product_info.xlsx"
with pd.ExcelWriter(file_nm) as writer:
    for file_name in file_list_xls:
        df = pd.read_excel("practice/참고자료/" + file_name) # Excel 파일의 데이터프레임 조회
        df.to_excel(writer, sheet_name = file_name.replace('.xlsx', '')) # 'product_info.xlsx' 파일에 Excel 파일의 이름으로 시트를 만들어 저장한다.
